"""生成 ledger-workbook.xlsx —— Temu/速卖通全托管场景财务大表模板。"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).resolve().parent / "ledger-workbook.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SAMPLE_FILL = PatternFill("solid", fgColor="FFF2CC")
NOTE_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def style_header(ws, headers, widths):
    ws.append(headers)
    for col, width in enumerate(widths, start=1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def add_sample(ws, rows):
    for row in rows:
        ws.append(row)
        for col in range(1, len(row) + 1):
            cell = ws.cell(ws.max_row, col)
            cell.fill = SAMPLE_FILL
            cell.border = THIN
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def add_validation(ws, col_letter, formula, end_row=800):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "请从下拉列表中选择"
    dv.errorTitle = "输入有误"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{end_row}")


def sheet_guide(wb):
    ws = wb.active
    ws.title = "使用说明"
    ws.column_dimensions["A"].width = 102
    lines = [
        "财务大表模板（Temu / 速卖通全托管为主）",
        "",
        "本文件按您现有账本语言设计：提货、出库、退供、提现、支出、工资与代发。",
        "黄色行是示例，看懂后可删，换成真实数据。",
        "",
        "全托管记住一句话：出库不等于卖了；提现才接近收到货款；退供单独记账。",
        "",
        "建议顺序：",
        "1）先填商品档案（货号、几件装、尺码、大码是否加价）",
        "2）搬最近一个月的提货 / 出库 / 退供 / 提现 / 支出 / 工资",
        "3）对照 mapping-guide.md 的对号入座表",
        "4）缺字段或叫法不同，写在备注里发回",
        "",
        "计量提醒：包数 × 件装数 = 件数；大码常另价；袋子规格和袋子费计入提货成本。",
        "一行建议：一个货号 + 一个尺码 + 一次业务。您习惯的尺码横排矩阵可先记草稿再展开。",
    ]
    for i, line in enumerate(lines, start=1):
        cell = ws.cell(i, 1, line)
        if i == 1:
            cell.font = Font(bold=True, size=14, color="1F4E79")
            cell.fill = NOTE_FILL
        elif line.startswith("建议") or line.startswith("全托管"):
            cell.font = Font(bold=True, size=11)


def sheet_sku(wb):
    ws = wb.create_sheet("商品档案")
    headers = [
        "货号", "品名", "品类", "件装数", "计量单位", "尺码",
        "默认单价", "大码是否加价", "大码单价", "主要平台", "是否在售", "备注",
    ]
    widths = [10, 22, 10, 8, 8, 8, 10, 12, 10, 12, 10, 24]
    style_header(ws, headers, widths)
    add_sample(
        ws,
        [
            ["3031", "保暖棉三角", "内裤", 6, "条", "M", 3.6, "是", 4.0, "Temu", "是", "示例可删"],
            ["3031", "保暖棉三角", "内裤", 6, "条", "L", 3.6, "是", 4.0, "Temu", "是", ""],
            ["3031", "保暖棉三角", "内裤", 6, "条", "2XL", 4.0, "是", 4.0, "Temu", "是", "大码价"],
            ["2511", "拉筒冰丝", "内裤", 6, "条", "M", 3.6, "否", "", "速卖通", "是", ""],
            ["XO四件套", "XO四件套", "套装", 4, "套", "L", 3.6, "否", "", "代发", "是", "代发常见"],
        ],
    )
    add_validation(ws, "C", '"内裤,文胸,套装,塑身,配件,其他"')
    add_validation(ws, "H", '"是,否"')
    add_validation(ws, "J", '"Temu,速卖通,TK自营,多平台,其他"')
    add_validation(ws, "K", '"是,否"')


def sheet_purchase(wb):
    ws = wb.create_sheet("提货明细")
    headers = [
        "日期", "单据号", "合作厂", "货号", "品名", "尺码",
        "包数", "件装数", "件数", "单位", "单价", "货款金额",
        "袋子规格", "袋子单价", "袋子数量", "袋子金额", "合计金额",
        "付款状态", "入库地点", "原始凭证", "备注",
    ]
    widths = [
        12, 12, 12, 10, 16, 8,
        8, 8, 8, 6, 8, 10,
        10, 8, 8, 8, 10,
        10, 10, 16, 18,
    ]
    style_header(ws, headers, widths)
    add_sample(
        ws,
        [
            [
                "2026-08-03", "26109739", "兴发厂", "3031", "保暖棉三角", "XL",
                5, 6, 30, "条", 4.0, 120,
                "20*30", 0.145, 5, 0.73, 120.73,
                "未付", "国内仓", "供应商手写单.jpg", "示例；手写单常无金额",
            ],
            [
                "2026-08-03", "26109739", "兴发厂", "3031", "保暖棉三角", "2XL",
                12, 6, 72, "条", 4.0, 288,
                "20*30", 0.145, 12, 1.74, 289.74,
                "未付", "国内仓", "供应商手写单.jpg", "同单分码分行",
            ],
            [
                "2026-08-03", "26109739", "兴发厂", "2511", "拉筒冰丝", "M",
                8, 6, 48, "条", 3.6, 172.8,
                "20*30", 0.145, 8, 1.16, 173.96,
                "未付", "国内仓", "供应商手写单.jpg", "",
            ],
        ],
    )
    add_validation(ws, "R", '"未付,部分付,已付"')


def sheet_outbound(wb):
    ws = wb.create_sheet("出库明细")
    headers = [
        "日期", "出库单号", "出库类型", "平台或对象", "货号", "品名", "尺码",
        "包数", "件装数", "件数", "物流单号", "运费",
        "关联代发或货件", "原始凭证", "备注",
    ]
    widths = [12, 12, 14, 12, 10, 16, 8, 8, 8, 8, 14, 8, 14, 16, 18]
    style_header(ws, headers, widths)
    add_sample(
        ws,
        [
            [
                "2026-08-09", "AE-0809", "速卖通送仓", "速卖通", "3031", "保暖棉三角", "M",
                10, 6, 60, "", 0, "", "8月速卖通.xlsx", "示例可删；送仓不确认收入",
            ],
            [
                "2026-08-10", "TM-0810", "Temu送仓", "Temu", "2511", "拉筒冰丝", "L",
                20, 6, 120, "YT123", 85, "货件#T778", "", "",
            ],
            [
                "2026-08-02", "DF-0802", "代发出库", "欣巧儿", "XO四件套", "L",
                2, 4, 8, "", 0, "8月欣巧儿代发", "出库退供表.png", "代发结算另见工资与代发",
            ],
            [
                "2026-08-11", "TK-0811", "TK自营出库", "TK自营", "2429", "抽针三角", "S",
                5, 1, 5, "SF999", 12, "订单#8821", "", "自营才接近销售出库",
            ],
        ],
    )
    add_validation(
        ws,
        "C",
        '"Temu送仓,速卖通送仓,TK自营出库,代发出库,调拨,其他"',
    )


def sheet_return(wb):
    ws = wb.create_sheet("退供明细")
    headers = [
        "日期", "退供单号", "平台", "货号", "品名", "尺码",
        "包数", "件装数", "件数", "退供原因", "是否可再售",
        "影响金额", "处理方式", "原始凭证", "备注",
    ]
    widths = [12, 12, 10, 10, 16, 8, 8, 8, 8, 12, 10, 10, 12, 16, 18]
    style_header(ws, headers, widths)
    add_sample(
        ws,
        [
            [
                "2026-08-15", "RT-TM-01", "Temu", "3031", "保暖棉三角", "2XL",
                3, 6, 18, "滞销退供", "是", 72, "回国内仓", "7月退供TEMU", "示例可删",
            ],
            [
                "2026-08-18", "RT-AE-02", "速卖通", "2511", "拉筒冰丝", "M",
                1, 6, 6, "质量问题", "否", 21.6, "报损", "7月速卖通退供", "",
            ],
        ],
    )
    add_validation(ws, "C", '"Temu,速卖通,TK自营,其他"')
    add_validation(ws, "J", '"滞销退供,质量问题,包装破损,少件多件,其他"')
    add_validation(ws, "K", '"是,否,待检"')
    add_validation(ws, "M", '"回国内仓,再送仓,转卖,报损,待定"')


def sheet_withdraw(wb):
    ws = wb.create_sheet("提现明细")
    headers = [
        "提现日期", "平台", "结算周期", "提现金额", "币种", "汇率", "人民币金额",
        "到账账户", "平台扣费合计", "扣费说明", "结算单号", "原始凭证", "备注",
    ]
    widths = [12, 10, 14, 12, 8, 8, 12, 12, 12, 20, 14, 16, 18]
    style_header(ws, headers, widths)
    add_sample(
        ws,
        [
            [
                "2026-08-25", "Temu", "2026-08-01~08-15", 8600, "CNY", 1, 8600,
                "对公账户", 420, "含售后扣款+活动扣费", "TM-SET-0815", "提现截图.png", "示例可删",
            ],
            [
                "2026-08-28", "速卖通", "2026-08-01~08-20", 5200, "CNY", 1, 5200,
                "支付宝", 180, "佣金类已内扣则备注", "AE-SET-0820", "", "",
            ],
        ],
    )
    add_validation(ws, "B", '"Temu,速卖通,TK自营,其他"')
    add_validation(ws, "E", '"CNY,USD,EUR,其他"')
    add_validation(ws, "H", '"对公账户,支付宝,微信,个人卡,其他"')


def sheet_expense(wb):
    ws = wb.create_sheet("支出明细")
    headers = [
        "日期", "类别", "对方", "金额", "账户", "关联单据", "原始凭证", "备注",
    ]
    widths = [12, 12, 14, 10, 12, 14, 16, 24]
    style_header(ws, headers, widths)
    add_sample(
        ws,
        [
            ["2026-08-03", "采购付款", "兴发厂", 5000, "微信", "26109739", "转账截图.jpg", "示例；对应提货"],
            ["2026-08-10", "运费物流", "货代", 85, "支付宝", "TM-0810", "运费账单.pdf", "送仓运费"],
            ["2026-08-31", "房租水电", "房东", 4500, "银行转账", "", "房租回单.jpg", ""],
            ["2026-08-31", "包材杂费", "包材店", 200, "微信", "", "", "零星袋子"],
        ],
    )
    add_validation(
        ws,
        "B",
        '"采购付款,运费物流,广告推广,房租水电,包材杂费,平台罚金,税费,其他"',
    )


def sheet_labor(wb):
    ws = wb.create_sheet("工资与代发")
    headers = [
        "日期", "对象类型", "姓名或厂名", "工序或项目", "货号或批次", "尺码",
        "数量", "数量单位", "单价", "金额", "附加费", "附加费说明",
        "结算状态", "结算日期", "备注",
    ]
    widths = [12, 10, 12, 12, 14, 8, 8, 8, 8, 8, 8, 14, 10, 12, 18]
    style_header(ws, headers, widths)
    add_sample(
        ws,
        [
            [
                "2026-08-19", "员工", "小陈", "质检", "3031", "M",
                200, "条", 0.15, 30, 0, "", "未结算", "", "示例可删",
            ],
            [
                "2026-08-19", "员工", "阿芳", "装袋", "3031", "混码",
                40, "包", 0.2, 8, 0, "", "未结算", "", "",
            ],
            [
                "2026-08-02", "代发厂", "欣巧儿", "代发加工", "XO四件套", "L",
                2, "包", 14.4, 28.8, 0.8, "包装+带子=0.4/包", "已结算", "2026-08-31",
                "对齐代发页算法",
            ],
        ],
    )
    add_validation(ws, "B", '"员工,代发厂,其他"')
    add_validation(ws, "D", '"分拣,质检,贴标,装袋,装箱,代发加工,返工,其他"')
    add_validation(ws, "H", '"条,包,套,件,箱"')
    add_validation(ws, "M", '"未结算,已结算,有争议"')


def main():
    wb = Workbook()
    sheet_guide(wb)
    sheet_sku(wb)
    sheet_purchase(wb)
    sheet_outbound(wb)
    sheet_return(wb)
    sheet_withdraw(wb)
    sheet_expense(wb)
    sheet_labor(wb)
    wb.save(OUT)
    print(f"已生成：{OUT}")
    print("SHEETS=", [s.encode("unicode_escape").decode() for s in wb.sheetnames])


if __name__ == "__main__":
    main()

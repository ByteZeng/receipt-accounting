"""spike 主入口：扫描样本目录 -> 识别 -> 结构校验 -> 输出 JSON 与标黄 Excel。

用法：
    python run_ocr.py                # 识别 samples/ 下全部图片
    python run_ocr.py --limit 5      # 先跑 5 张试水，确认 Key 和格式没问题
    python run_ocr.py --no-cache     # 忽略缓存重新识别（改了提示词后用）
"""

import argparse
import concurrent.futures as futures
import hashlib
import json
import sys
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

import config
import ocr_client
from normalize import normalize_amount, normalize_date

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

# qwen-vl-max 华北2 原价，2026-09-03 查自阿里云百炼文档，单位美元/百万 token。
# 仅用于量级估算，实际以账单为准。
PRICE_INPUT_USD = 0.229
PRICE_OUTPUT_USD = 0.573
USD_TO_CNY = 7.1

YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def list_images() -> List[Path]:
    if not config.SAMPLES_DIR.exists():
        raise SystemExit(f"样本目录不存在：{config.SAMPLES_DIR}")
    files = [
        p
        for p in sorted(config.SAMPLES_DIR.rglob("*"))
        if p.is_file() and p.suffix.lower() in config.IMAGE_SUFFIXES
    ]
    if not files:
        raise SystemExit(f"{config.SAMPLES_DIR} 下没有找到图片")
    return files


def field(raw: Dict[str, Any], name: str) -> Dict[str, Any]:
    value = raw.get(name) or {}
    if not isinstance(value, dict):
        value = {"value": value, "confidence": 0.0}
    return {
        "value": value.get("value"),
        "confidence": float(value.get("confidence") or 0.0),
    }


def cross_check(items: List[Dict[str, Any]], total: Decimal) -> str:
    """明细小计之和与合计金额比对。不一致往往意味着某个数字识别错了。"""
    if not items or total is None:
        return ""

    subtotal = Decimal("0")
    for item in items:
        amount, _ = normalize_amount(item.get("amount"))
        if amount is None:
            return ""  # 有明细读不出，无法比对，不误报
        subtotal += amount

    diff = abs(subtotal - total)
    if diff > Decimal(config.AMOUNT_TOLERANCE):
        return f"明细合计 {subtotal} 与单据金额 {total} 不符（差 {diff}），需人工核对"
    return ""


def process(path: Path, use_cache: bool) -> Dict[str, Any]:
    digest = file_hash(path)
    cache_file = config.CACHE_DIR / f"{digest}.json"

    if use_cache and cache_file.exists():
        raw = json.loads(cache_file.read_text(encoding="utf-8"))
        cached = True
    else:
        raw = ocr_client.extract(path)
        cache_file.write_text(
            json.dumps(raw, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        cached = False

    date_field = field(raw, "date")
    supplier_field = field(raw, "supplier")
    amount_field = field(raw, "total_amount")
    doc_no_field = field(raw, "doc_no")

    date_value, date_note = normalize_date(date_field["value"], config.FALLBACK_YEAR)
    amount_value, amount_note = normalize_amount(amount_field["value"])

    items = raw.get("line_items") or []
    notes = [n for n in (raw.get("notes"), date_note, amount_note) if n]
    check_note = cross_check(items if isinstance(items, list) else [], amount_value)
    if check_note:
        notes.append(check_note)

    confidences = [
        date_field["confidence"],
        supplier_field["confidence"],
        amount_field["confidence"],
    ]
    min_confidence = min(confidences)

    needs_review = bool(
        min_confidence < config.CONFIDENCE_THRESHOLD
        or raw.get("illegible")
        or check_note
        or amount_value is None
        or date_value is None
    )

    usage = raw.get("_usage") or {}
    return {
        "file": str(path.relative_to(config.SAMPLES_DIR)),
        "hash": digest,
        "doc_type": raw.get("doc_type") or "无法判断",
        "date": date_value,
        "date_raw": date_field["value"],
        "date_confidence": date_field["confidence"],
        "supplier": supplier_field["value"],
        "supplier_confidence": supplier_field["confidence"],
        "doc_no": doc_no_field["value"],
        "amount": str(amount_value) if amount_value is not None else None,
        "amount_raw": amount_field["value"],
        "amount_confidence": amount_field["confidence"],
        "line_item_count": len(items) if isinstance(items, list) else 0,
        "illegible": bool(raw.get("illegible")),
        "needs_review": needs_review,
        "notes": "；".join(notes),
        "cached": cached,
        "prompt_tokens": usage.get("prompt_tokens", 0),
        "completion_tokens": usage.get("completion_tokens", 0),
    }


HEADERS = [
    ("file", "原始文件", 28),
    ("doc_type", "单据类型", 12),
    ("date", "日期", 12),
    ("supplier", "供应商", 20),
    ("doc_no", "单据号", 14),
    ("amount", "金额", 12),
    ("line_item_count", "明细行数", 10),
    ("needs_review", "需复核", 9),
    ("notes", "备注与异常", 50),
]


def write_excel(rows: List[Dict[str, Any]], errors: List[Dict[str, str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "识别明细"

    for column, (_, title, width) in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=column, value=title)
        cell.font = Font(bold=True)
        sheet.column_dimensions[cell.column_letter].width = width
    sheet.freeze_panes = "A2"

    confidence_map = {
        "date": "date_confidence",
        "supplier": "supplier_confidence",
        "amount": "amount_confidence",
    }

    for index, row in enumerate(rows, start=2):
        for column, (key, _, _) in enumerate(HEADERS, start=1):
            value = row.get(key)
            if key == "needs_review":
                value = "是" if value else ""
            cell = sheet.cell(row=index, column=column, value=value)

            # 标黄的含义是「模型自己承认不确定」，客户只需复核这些格子
            conf_key = confidence_map.get(key)
            if conf_key and row.get(conf_key, 1.0) < config.CONFIDENCE_THRESHOLD:
                cell.fill = YELLOW
            if key == "notes" and value:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    summary = workbook.create_sheet("按供应商汇总")
    summary.append(["供应商", "单据数", "金额合计", "含需复核"])
    for cell in summary[1]:
        cell.font = Font(bold=True)
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["C"].width = 14
    summary.column_dimensions["D"].width = 12

    grouped: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        key = row.get("supplier") or "未识别"
        bucket = grouped.setdefault(key, {"count": 0, "total": Decimal("0"), "review": 0})
        bucket["count"] += 1
        if row.get("amount"):
            bucket["total"] += Decimal(row["amount"])
        if row.get("needs_review"):
            bucket["review"] += 1

    for name, bucket in sorted(grouped.items(), key=lambda kv: -kv[1]["total"]):
        summary.append([name, bucket["count"], float(bucket["total"]), bucket["review"]])

    if errors:
        sheet_errors = workbook.create_sheet("识别失败")
        sheet_errors.append(["文件", "失败原因"])
        for cell in sheet_errors[1]:
            cell.font = Font(bold=True)
        sheet_errors.column_dimensions["A"].width = 28
        sheet_errors.column_dimensions["B"].width = 60
        for item in errors:
            sheet_errors.append([item["file"], item["error"]])
            sheet_errors.cell(row=sheet_errors.max_row, column=1).fill = RED

    workbook.save(config.RESULTS_XLSX)


def report(rows: List[Dict[str, Any]], errors: List[Dict[str, str]]) -> None:
    total = len(rows)
    review = sum(1 for r in rows if r["needs_review"])
    prompt_tokens = sum(r["prompt_tokens"] for r in rows)
    completion_tokens = sum(r["completion_tokens"] for r in rows)
    fresh = sum(1 for r in rows if not r["cached"])

    cost = (
        prompt_tokens / 1e6 * PRICE_INPUT_USD
        + completion_tokens / 1e6 * PRICE_OUTPUT_USD
    ) * USD_TO_CNY

    print()
    print("=" * 56)
    print(f"识别完成 {total} 张，失败 {len(errors)} 张，本次实际调用 {fresh} 张")
    print(f"需人工复核 {review} 张（{review / total:.0%}）" if total else "")
    if fresh:
        print(f"本次 token：输入 {prompt_tokens}，输出 {completion_tokens}")
        print(f"估算费用：{cost:.4f} 元，折合每张 {cost / fresh:.4f} 元")
        print(f"按月 300 张推算：约 {cost / fresh * 300:.2f} 元/月")
    print(f"结果 JSON：{config.RESULTS_JSON}")
    print(f"结果表格：{config.RESULTS_XLSX}")
    print("=" * 56)
    print("下一步：填好 ground_truth.csv 后运行 python eval_accuracy.py")


def main() -> None:
    parser = argparse.ArgumentParser(description="单据识别 spike")
    parser.add_argument("--limit", type=int, default=0, help="只处理前 N 张")
    parser.add_argument("--no-cache", action="store_true", help="忽略缓存重新识别")
    args = parser.parse_args()

    config.ensure_dirs()
    images = list_images()
    if args.limit:
        images = images[: args.limit]

    print(f"待处理 {len(images)} 张，模型 {config.MODEL}，并发 {config.MAX_WORKERS}")

    rows: List[Dict[str, Any]] = []
    errors: List[Dict[str, str]] = []

    with futures.ThreadPoolExecutor(max_workers=config.MAX_WORKERS) as pool:
        tasks = {
            pool.submit(process, path, not args.no_cache): path for path in images
        }
        for done, task in enumerate(futures.as_completed(tasks), start=1):
            path = tasks[task]
            name = path.name
            try:
                rows.append(task.result())
                print(f"[{done}/{len(images)}] {name}")
            except Exception as exc:
                errors.append({"file": name, "error": str(exc)})
                print(f"[{done}/{len(images)}] {name} 失败：{exc}")

    rows.sort(key=lambda r: (r["date"] or "9999", r["file"]))
    config.RESULTS_JSON.write_text(
        json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    write_excel(rows, errors)
    report(rows, errors)


if __name__ == "__main__":
    main()
